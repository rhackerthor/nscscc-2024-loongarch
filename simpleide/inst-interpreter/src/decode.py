from openpyxl import load_workbook

class InstDecoder:
  def __init__(self, filename='inst.xlsx'):
    # 打开表格
    self.wb = load_workbook(filename)
    self.ws = self.wb[self.wb.sheetnames[0]]
    self.rows = self.ws.max_row
    self.beginrow = 2

    self.name_line = 1
    self.value_line = 2
    self.enable_line = 3
    self.type_line = 4
    self.rf_we_line = 5
    self.rf_oe1_line = 6
    self.rf_oe2_line = 7
    self.is_rd_line = 8
    self.sel_alu_in1_line = 9
    self.sel_alu_in2_line = 10

    self.format_sheet()
    self.save()

  def format_sheet(self):
    # 格式化表格
    for i in range(self.beginrow, self.rows + 1):
      inst_name = self.ws.cell(i, 1).value
      if inst_name and isinstance(inst_name, str):
        formatted_name = '_' + inst_name.lower().replace("_", "", 1)
        self.ws.cell(row=i, column=1, value=formatted_name)

      inst_value = self.ws.cell(i, 2).value
      if inst_value and isinstance(inst_value, str):
        formatted_value = inst_value.replace(" ", "")
        self.ws.cell(row=i, column=2, value=formatted_value)

  def save(self, filename='inst.xlsx'):
    # 保存表格
    self.wb.save(filename)

  def cell(self, row: int, column: int):
    return self.ws.cell(row, column).value

  def get_name(self, i) -> str:
    return str(self.cell(i, self.name_line))
  
  def get_value(self, i) -> str:
    return str(self.cell(i, self.value_line))

  def get_type(self, i) -> str:
    return str(self.cell(i, self.type_line))

  def get_enable(self, i) -> int:
    return int(self.cell(i, self.enable_line))

  def generate_verilog_signal(self, file_name: str):
    """生成verilog代码所需的译码信号"""
    with open(file_name, 'w') as fout:
      # 定义信号
      for i in range(self.beginrow, self.rows + 1):
        if self.get_enable(i):
          fout.write(f"logic {self.get_name(i)};\n")

      fout.write("\n")

      # 信号赋值
      for i in range(self.beginrow, self.rows + 1):
        if self.get_enable(i):
          fout.write(f"assign U_D.{self.get_name(i):<10} = ")
          value_str = self.get_value(i)
          opcode = [value_str[j:j + 6] for j in range(0, len(value_str), 6)]
          conditions = []
          start = 31

          for subop in opcode:
            if len(subop) == 1:
              condition = f"(inst[{start}] == 1'b{subop})"
            else:
              condition = f"(inst[{start}:{start - len(subop) + 1}] == {len(subop)}'b{subop})"
            conditions.append(condition)
            start -= len(subop)

          fout.write(" && ".join(conditions) + ";\n")

  def generate_emulator_signal(self, file_name: str):
    """生成c++仿真器所需的译码语句"""
    with open(file_name, "w") as fout:
      for i in range(self.beginrow, self.rows + 1):
        name = self.get_name(i).replace("_", "")
        inst = self.get_value(i) + '?' * (32 - len(self.get_value(i)))
        type_ = self.get_type(i)

        if self.get_enable(i) == 1:
          fout.write(f"INSTPAT(\"{inst}\", {name:<10}, {type_:<5}, );\n")

  def generate_verilog_ctrl_signal(self, title: str, dir_path: str, column: int):
    """生成cpu控制信号"""
    with open(f"{dir_path}/{title}.txt", "w") as fout:
      # 将行转化成列表
      value_list = [self.cell(i, column) for i in range(self.beginrow, self.rows + 1)]

      # 将列表转换成字典
      value_to_indices = {}
      for index, value in enumerate(value_list):
        if value not in value_to_indices:
          value_to_indices[value] = []
        if self.get_enable(index + 2) == 1:
          value_to_indices[value].append(f"U_D.{self.get_name(index + 2)}")

      # 输出
      fout.write(f"{title}:\n")
      for key, value in value_to_indices.items():
        fout.write(f"{key}: {len(value)}\n")
        for i in range(0, len(value), 4):
          fout.write(", ".join(value[i:i+4]) + ",\n")
        fout.write("\n")
      