from openpyxl import load_workbook
# import pandas as pd
# import copy

class inst_decoder(object):
  def __init__(self):
    # 打开表格
    wb = load_workbook('inst.xlsx')
    self.ws = wb[wb.sheetnames[0]]
    self.rows = self.ws.max_row
    # 格式化表格
    for i in range(1, self.rows + 1):
      inst_name = self.ws.cell(i, 1).value
      if inst_name is not None and isinstance(inst_name, str):
        self.ws.cell(row=i, column=1, value='inst_' + inst_name.lower().replace(".", "_").replace("inst_", ""))

      inst_value = self.ws.cell(i, 2).value
      if inst_value is not None and isinstance(inst_value, str):
        self.ws.cell(row=i, column=2, value=inst_value.replace(" ", ""))
    # 保存
    wb.save('inst.xlsx')
  
  def cell(self, c, r):
    return self.ws.cell(c, r).value
  
  def decode(self, file_name):
    with open(file_name, 'w') as fout:
      # 定义信号
      for i in range(1, self.rows + 1):
        if int(self.cell(i, 3)) == 0:
          continue
        fout.write("logic %s;\n" % self.cell(i, 1))
      fout.write("\n")
      # 信号赋值
      for i in range(1, self.rows + 1):
        if int(self.cell(i, 3)) == 0:
          continue
        fout.write("assign %-14s = "  % self.cell(i, 1))
        s = self.cell(i, 2)
        opcode = [s[j:j + 6] for j in range(0, len(s), 6)]
        flag = True
        start = 31
        for subop in opcode:
          if flag is False:
            fout.write(" & ")
          if len(subop) == 1:
            if subop[0] == "0":
              fout.write("~")
            fout.write("inst_i[%d]" % start)
          else:
            fout.write("opcode_%d_%d[%d'b%s]" % (start, start - len(subop) + 1, len(subop), subop))
          start -= len(subop)
          flag = False
        fout.write(";\n")

  def instpat(self, file_name):
    with open(file_name, "w") as fout:
      for i in range(1, self.rows + 1):
        name = self.cell(i, 1)
        inst = self.cell(i, 2)
        type = self.cell(i, 4)

        name = name.replace("inst_", "")
        inst += '?' * (32 - len(inst))
        if int(self.cell(i, 3)) == 1: 
          fout.write("INSTPAT(\"%s\", %-10s, %-5s, );\n" %(inst, name, type))


def main():
  D = inst_decoder()
  D.decode("build/decode.txt")
  D.instpat("build/instpat.txt")
  
if __name__=="__main__":
  main()